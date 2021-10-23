import tkinter
from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl import *
from Human_Resource_Emp import homescreen

#Screen that lets employee choose - Register and login 
def first_screen():
    global root1
    root1 = Tk()
    root1.title('Employee Choose')
    root1.geometry('700x700')
    root1.config(bg = "gainsboro")

    labelemp = Label(root1, text = "Welcome Employee!",fg = 'purple4', bg = "gainsboro", font = ("Times","25","bold")).\
            place(x="230", y="50")

    labelopt = Label(root1, text = "Select an Option",fg = 'purple4', bg = "gainsboro", font = ("Times","20")).\
            place(x="280", y="130")

    signupbutton = Button(root1, text='Sign Up', bg = 'purple4', fg = 'white', font = ('Times','20'),command = signup)
    signupbutton.place(x = "300", y = "200", width = '200', height = '60')

    signinbutton = Button(root1, text='Log In', bg = 'purple4', fg = 'white', font = ('Times','20'),command = empscreen)
    signinbutton.place(x = "300", y = "300", width = '200', height = '60')
    
def signup():
    global sign_win
    sign_win = Tk()
    sign_win.title('Employee Signup')
    sign_win.geometry('700x700')
    sign_win.config(bg = "gainsboro")

    labelsign = Label(sign_win, text = "Enter a new Username and Password:",fg = 'purple4', bg = "gainsboro", \
                      font = ("Times","20",)).place(x="170", y="120")

    
    labeluser = Label(sign_win, text = "Username *", bg = "gainsboro", font = ("Times","14")).\
            place(x="200", y="200")

    global userentry2, passentry2
    
    userentry2 = Entry(sign_win, font =("Times","13"))
    userentry2.place(x = "200", y = "240", width = "300", height = "30")

    passlabel = Label(sign_win, text = "Password *", bg = "gainsboro", font = ("Times","14")).\
            place(x="200", y="300")

    passentry2 = Entry(sign_win, font =("Times","13"), show = '*')
    passentry2.place(x = "200", y = "340", width = "300", height = "30")

    global signbutton
    
    signbutton = Button(sign_win, text='Sign Up', bg = 'purple4', fg = 'white', font = ('15'),command = register)
    signbutton.place(x = "280", y = "400", width = '100', height = '30')
    sign_win.mainloop()
    return root1.destroy()

#Employee Login   
def empscreen():
    root1.destroy()
    global root
    root = Tk()
    root.title('Employee Login')
    root.geometry('700x700')
    root.config(bg = "gainsboro")

    labeladmin = Label(root, text = "Welcome Employee!",fg = 'purple4', bg = "gainsboro", font = ("Times","25","bold")).\
            place(x="230", y="50")
    labelsign = Label(root, text = "Sign In",fg = 'purple4', bg = "gainsboro", font = ("Times","20")).\
            place(x="300", y="130")

    labeluser = Label(root, text = "Username *", bg = "gainsboro", font = ("Times","14")).\
            place(x="200", y="200")

    global userentry3, passentry3
    
    userentry3 = Entry(root, font =("Times","13"))
    userentry3.place(x = "200", y = "240", width = "300", height = "30")

    passlabel = Label(root, text = "Password *", bg = "gainsboro", font = ("Times","14")).\
            place(x="200", y="300")

    passentry3 = Entry(root, font =("Times","13"),show = '*')
    passentry3.place(x = "200", y = "340", width = "300", height = "30")

    global signbutton3

    signbutton3 = Button(root, text='Sign In', bg = 'purple4', fg = 'white', font = ('15'), command = verify)
    signbutton3.place(x = "280", y = "400", width = '100', height = '30')

    root.mainloop()
    return 

def register():
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["regis"]
    list1 = []
    list2 = []
    for i in range(1,ws.max_row):
        list1.append(ws[i][0].value)
        list2.append(ws[i][1].value)


    data = [[userentry2.get(), passentry2.get()]]
    
    if userentry2.get() not in list1:
        if len(passentry2.get()) >= 8:
            for i in data:
                ws.append(i)
                messagebox.showinfo("Success","Signup Successful!")
                wb.save("HRM_Excel.xlsx")
            return sign_win.destroy()
                

        else:
            messagebox.showinfo("Weak Password","Password should be greater than 8 characters")
            
    elif userentry2.get() in list1:
        messagebox.showinfo("Invalid Username","Username taken!")
        
        
    wb.save("HRM_Excel.xlsx")

    userentry2.delete(0, END)
    passentry2.delete(0, END)
    

    

def verify():
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["regis"]
    list1 = []
    list2 = []
    for i in range(1,ws.max_row):
        list1.append(ws[i][0].value)
        list2.append(ws[i][1].value)

    if userentry3.get() in list1 and passentry3.get() in list2:
        signbutton3.config(command = homescreen)
        return root.destroy(),homescreen()

    else:
        messagebox.showinfo("Not Recognized","Username or password not recognized, Try again!")

        root.mainloop()
        

