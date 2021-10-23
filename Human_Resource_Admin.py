import tkinter
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import tkinter.simpledialog
from PIL import ImageTk, Image
import openpyxl
from openpyxl import *
import xlrd
import xlwt
from tkinter.ttk import *
import random

from datetime import date

def profileNew():
    global win_count
    win_count+=1
    if win_count >1:
        messagebox.showinfo("warning","Window is already opened")
        return
    
    def cancel():
        global win_count
        win_count = 0
        addwindow.destroy()
        
    global addwindow,gradentry,designationentry
    addwindow = tkinter.Tk()
    addwindow.title("Add New Employee")
    addwindow.resizable(False, False)
    canvas = tkinter.Canvas(addwindow, height=700, width=900) 

    frame = tkinter.Frame(addwindow, bg='#5F9EA0')
    frame.place(relwidth=1, relheight=1)

    boxlabel = tkinter.Label(frame, text='Main Profile', bg='#5F9EA0', fg='#FFFFFF')
    boxlabel.place(relx=0.05, rely=0.025)

    mainbox = tkinter.Canvas(frame, height=100, width=800, bd=2, bg='#5F9EA0')# the box of main profile
    mainbox.place(relx=0.05, rely=0.05)

    global idgen,namentry,typentry,idgenlabel
    idlabel = tkinter.Label(mainbox, text="Employee ID", bg='#5F9EA0', fg='#FFFFFF')
    idlabel.place(relx=0.03, rely=0.2, relwidth=0.14)

    #Randomly generate ID
    idgen = 'E' + str(random.randint(100,500))
    
    idgenlabel = tkinter.Label(mainbox, text= idgen, bg='#5F9EA0', fg='white')
    idgenlabel.place(relx=0.19, rely=0.2, relwidth=0.2)

    empnamelabel = tkinter.Label(mainbox, text="Employee Name", bg='#5F9EA0', fg='#FFFFFF')
    empnamelabel.place(relx=0.03, rely=0.6, relwidth=0.14)

    empnamentry = tkinter.Entry(mainbox)
    empnamentry.place(relx=0.19, rely=0.6, relwidth=0.2)

    designationlabel = tkinter.Label(mainbox, text="Position", bg='#5F9EA0', fg='#FFFFFF')
    designationlabel.place(relx=0.64, rely=0.2, relwidth=0.1)

    designationentry = tkinter.Entry(mainbox)
    designationentry.place(relx=0.765, rely=0.2, relwidth=0.2)

    personalboxlabel = tkinter.Label(frame, text='Personal Profile', bg='#5F9EA0', fg='#FFFFFF')
    personalboxlabel.place(relx=0.05, rely=0.225)

    personalbox = tkinter.Canvas(frame,height=300, width=800, bd=2, bg='#5F9EA0')# the box for personal profile
    personalbox.place(relx=0.05, rely=0.25)

    namelabel = tkinter.Label(personalbox, text="First Name", bg='#5F9EA0', fg='#FFFFFF')
    namelabel.place(relx=0.03, rely=0.1, relwidth=0.1)

    namentry = tkinter.Entry(personalbox)
    namentry.place(relx=0.15, rely=0.1, relwidth=0.2)

    sexlabel = tkinter.Label(personalbox, text="Sex", bg='#5F9EA0', fg='#FFFFFF')
    sexlabel.place(relx=0.03, rely=0.2, relwidth=0.1)

    global sexcombo
    sexlist = ["Male","Female"]
    sexcombo = ttk.Combobox(personalbox, values = sexlist)
    sexcombo.place(relx=0.15, rely=0.2, relwidth=0.1)

    global bloodcombo,placentry, datecombo, monthcombo, yearcombo
    bloodlabel = tkinter.Label(personalbox, text="Blood Group", bg='#5F9EA0', fg='#FFFFFF')
    bloodlabel.place(relx=0.03, rely=0.3, relwidth=0.1)

    bloodlist = ["A+ve","B+ve","O+ve","AB","A-ve","B-ve","O-ve"]
    bloodcombo = ttk.Combobox(personalbox, values = bloodlist)
    bloodcombo.place(relx=0.15, rely=0.3, relwidth=0.09)

    datelabel = tkinter.Label(personalbox, text="Date of birth", bg='#5F9EA0', fg='#FFFFFF')
    datelabel.place(relx=0.03, rely=0.4, relwidth=0.1)

    datelabel = tkinter.Label(personalbox, text="Month:", bg='#5F9EA0', fg='#FFFFFF')
    datelabel.place(relx=0.15, rely=0.4, relwidth=0.075)

    monthlist = [i for i in range(1,13)]
    monthcombo = ttk.Combobox(personalbox, values = monthlist, width = 2)
    monthcombo.place(relx=0.22, rely=0.4, relwidth=0.075)

    datelabel = tkinter.Label(personalbox, text="Day:", bg='#5F9EA0', fg='#FFFFFF')
    datelabel.place(relx=0.3, rely=0.4, relwidth=0.075)

    datelist = [i for i in range (1,32)]
    datecombo = ttk.Combobox(personalbox, values = datelist, width = 2)
    datecombo.place(relx=0.38, rely=0.4, relwidth=0.075)

    datelabel = tkinter.Label(personalbox, text="Year:", bg='#5F9EA0', fg='#FFFFFF')
    datelabel.place(relx=0.46, rely=0.4, relwidth=0.075)

    yearlist = [i for i in range(1959,2001)]
    yearcombo = ttk.Combobox(personalbox, values = yearlist, width = 2)
    yearcombo.place(relx=0.54, rely=0.4, relwidth=0.075)
    
    #datentry.place(relx=0.15, rely=0.4, relwidth=0.2)

    placelabel = tkinter.Label(personalbox, text="Place of birth", bg='#5F9EA0', fg='#FFFFFF')
    placelabel.place(relx=0.03, rely=0.5, relwidth=0.1)

    placentry = tkinter.Entry(personalbox)
    placentry.place(relx=0.15, rely=0.5, relwidth=0.2)

    nationlabel = tkinter.Label(personalbox, text="Nationality", bg='#5F9EA0', fg='#FFFFFF')
    nationlabel.place(relx=0.03, rely=0.6, relwidth=0.1)

    global nationentry,townentry,housentry
    nationentry = tkinter.Entry(personalbox)
    nationentry.place(relx=0.15, rely=0.6, relwidth=0.2)

    townlabel = tkinter.Label(personalbox, text="Town", bg='#5F9EA0', fg='#FFFFFF')
    townlabel.place(relx=0.03, rely=0.7, relwidth=0.1)

    townentry = tkinter.Entry(personalbox)
    townentry.place(relx=0.15, rely=0.7, relwidth=0.2)

    houselabel = tkinter.Label(personalbox, text="House Number", bg='#5F9EA0', fg='#FFFFFF')
    houselabel.place(relx=0.03, rely=0.8, relwidth=0.1)

    housentry = tkinter.Entry(personalbox)
    housentry.place(relx=0.15, rely=0.8, relwidth=0.2)
    
    global mobilentry,telentry,emailentry,departmententry
    mobilelabel = tkinter.Label(personalbox, text="Mobile Number", bg='#5F9EA0', fg='#FFFFFF')
    mobilelabel.place(relx=0.028, rely=0.9, relwidth=0.12)

    mobilentry = tkinter.Entry(personalbox)
    mobilentry.place(relx=0.15, rely=0.9, relwidth=0.2)

    telelabel = tkinter.Label(personalbox, text="Telephone Number", bg='#5F9EA0', fg='#FFFFFF')
    telelabel.place(relx=0.61, rely=0.8, relwidth=0.13)

    telentry = tkinter.Entry(personalbox)
    telentry.place(relx=0.765, rely=0.8, relwidth=0.2)

    emailabel = tkinter.Label(personalbox, text="E-mail", bg='#5F9EA0', fg='#FFFFFF')
    emailabel.place(relx=0.64, rely=0.9, relwidth=0.1)

    emailentry = tkinter.Entry(personalbox)
    emailentry.place(relx=0.765, rely=0.9, relwidth=0.2)

    joboxlabel = tkinter.Label(frame, text='Job detail', bg='#5F9EA0', fg='#FFFFFF')
    joboxlabel.place(relx=0.05, rely=0.7)

    jobox = tkinter.Canvas(frame,height=150, width=800, bd=2, bg='#5F9EA0')# box for job details
    jobox.place(relx=0.05, rely=0.725)

    departmentlabel = tkinter.Label(jobox, text="Department", bg='#5F9EA0', fg='#FFFFFF')
    departmentlabel.place(relx=0.05, rely=0.125, relwidth=0.1)

    departmententry = tkinter.Entry(jobox)
    departmententry.place(relx=0.15, rely=0.125, relwidth=0.2)

    global joinentry, confirmentry, lastentry, salaryentry, perentry
    
    joinlabel = tkinter.Label(jobox, text="Date of Joining", bg='#5F9EA0', fg='#FFFFFF')
    joinlabel.place(relx=0.64, rely=0.125, relwidth=0.1)

    joinentry = tkinter.Entry(jobox)
    joinentry.place(relx=0.765, rely=0.125, relwidth=0.2)

    confirmlabel = tkinter.Label(jobox, text="Date of confirmation", bg='#5F9EA0', fg='#FFFFFF')
    confirmlabel.place(relx=0.6, rely=0.375, relwidth=0.14)

    confirmentry = tkinter.Entry(jobox)
    confirmentry.place(relx=0.765, rely=0.375, relwidth=0.2)

    lastlabel = tkinter.Label(jobox, text="Date of last Increment", bg='#5F9EA0', fg='#FFFFFF')
    lastlabel.place(relx=0.6, rely=0.625, relwidth=0.14)

    lastentry = tkinter.Entry(jobox)
    lastentry.place(relx=0.765, rely=0.625, relwidth=0.2)

    salarylabel = tkinter.Label(jobox, text="Salary", bg='#5F9EA0', fg='#FFFFFF')
    salarylabel.place(relx=0.03, rely=0.375, relwidth=0.14)

    salaryentry = tkinter.Entry(jobox)
    salaryentry.place(relx=0.15, rely=0.375, relwidth=0.2)

    perlabel = tkinter.Label(jobox, text="Salary per day", bg='#5F9EA0', fg='#FFFFFF')
    perlabel.place(relx=0.03, rely=0.625, relwidth=0.14)

    perentry = tkinter.Entry(jobox)
    perentry.place(relx=0.15, rely=0.625, relwidth=0.2)

    savebutton = tkinter.Button(frame, text='Save', command = getinput)
    savebutton.place(relx=0.15, rely=0.955, relwidth=0.1)

    cancelbutton = tkinter.Button(frame, text='Cancel', command = cancel)
    cancelbutton.place(relx=0.275, rely=0.955, relwidth=0.1)

    canvas.pack()

    addwindow.mainloop()
    return addwindow

def profileEdit():
    global win1_count
    win1_count+=1
    if win1_count >1:
        messagebox.showinfo("warning","Window is already opened")
        return
    def cancel():
        global win1_count
        win1_count = 0
        profwindow.destroy()
    excel=('HRM_Excel.xlsx')
    wb=xlrd.open_workbook(excel)
    sheet=wb.sheet_by_name("Employee Job Details")

    list1=[]                #list of id
    for i in range(2,sheet.nrows):
        list1.append(sheet.cell_value(i,0))

    global profwindow,gradentry,designationentry
    profwindow = tkinter.Tk()
    profwindow.title("Update employee's Personal Data")
    profwindow.resizable(False, False)
    canvas = tkinter.Canvas(profwindow, height=700, width=900) 

    frame = tkinter.Frame(profwindow, bg='#5F9EA0')
    frame.place(relwidth=1, relheight=1)

    boxlabel = tkinter.Label(frame, text='Main Profile', bg='#5F9EA0', fg='#FFFFFF')
    boxlabel.place(relx=0.05, rely=0.025)

    mainbox = tkinter.Canvas(frame, height=100, width=800, bd=2, bg='#5F9EA0')# the box of main profile
    mainbox.place(relx=0.05, rely=0.05)

    global idcombo,namentry,typentry,combo
    idlabel = tkinter.Label(mainbox, text="Employee ID", bg='#5F9EA0', fg='#FFFFFF')
    idlabel.place(relx=0.03, rely=0.2, relwidth=0.14)

    idcombo=Combobox(mainbox,values=list1)
    idcombo.place(relx=0.19, rely=0.2, relwidth=0.2)

    list2=["-","-","-","-","-","-","-","-","-"]

    def displayjob():
        list2.clear()
        for i in range(1,sheet.nrows): 
            if idcombo.get()==sheet.cell_value(i,0):

                list2.append(sheet.cell_value(i,1))
                designationview = tkinter.Label(mainbox,text=list2[0], bg='#5F9EA0', fg='#FFFFFF')
                designationview.place(relx=0.75, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,2))
                empnameview = tkinter.Label(mainbox,text=list2[1], bg='#5F9EA0', fg='#FFFFFF')
                empnameview.place(relx=0.19, rely=0.6, relwidth=0.2)

                list2.append(sheet.cell_value(i,3))
                departmentview = tkinter.Label(jobox, text=list2[2], bg='#5F9EA0', fg='#FFFFFF')
                departmentview.place(relx=0.15, rely=0.125, relwidth=0.2)

                list2.append(sheet.cell_value(i,4))
                salaryview = tkinter.Label(jobox, text=list2[3], bg='#5F9EA0', fg='#FFFFFF')
                salaryview.place(relx=0.15, rely=0.375, relwidth=0.2)

                
                list2.append(sheet.cell_value(i,5))
                perview = tkinter.Label(jobox, text=list2[4], bg='#5F9EA0', fg='#FFFFFF')
                perview.place(relx=0.15, rely=0.625, relwidth=0.2)


             
                list2.append(sheet.cell_value(i,6))
                joinview = tkinter.Label(jobox, text=list2[5], bg='#5F9EA0', fg='#FFFFFF')
                joinview.place(relx=0.765, rely=0.125, relwidth=0.2)

                
                list2.append(sheet.cell_value(i,7))
                confirmview = tkinter.Label(jobox, text=list2[6], bg='#5F9EA0', fg='#FFFFFF')
                confirmview.place(relx=0.765, rely=0.375, relwidth=0.2)

               
                list2.append(sheet.cell_value(i,8))
                lastview = tkinter.Label(jobox, text=list2[7], bg='#5F9EA0', fg='#FFFFFF')
                lastview.place(relx=0.765, rely=0.625, relwidth=0.2)

            
                list2.append(sheet.cell_value(i,9))
                            


    designationview = tkinter.Label(mainbox,text=list2[0], bg='#5F9EA0', fg='#FFFFFF')
    designationview.place(relx=0.19, rely=0.6, relwidth=0.2)

    designationlabel = tkinter.Label(mainbox, text="Employee Name", bg='#5F9EA0', fg='#FFFFFF')
    designationlabel.place(relx=0.03, rely=0.6, relwidth=0.14)

    empnamelabel = tkinter.Label(mainbox, text="Position", bg='#5F9EA0', fg='#FFFFFF')
    empnamelabel.place(relx=0.6, rely=0.2, relwidth=0.1)

    empnameview = tkinter.Label(mainbox,text=list2[1], bg='#5F9EA0', fg='#FFFFFF')
    empnameview.place(relx=0.75, rely=0.2, relwidth=0.2)

    personalboxlabel = tkinter.Label(frame, text='Personal Profile', bg='#5F9EA0', fg='#FFFFFF')
    personalboxlabel.place(relx=0.05, rely=0.225)

    personalbox = tkinter.Canvas(frame,height=300, width=800, bd=2, bg='#5F9EA0')# the box for personal profile
    personalbox.place(relx=0.05, rely=0.25)

    namelabel = tkinter.Label(personalbox, text="First Name", bg='#5F9EA0', fg='#FFFFFF')
    namelabel.place(relx=0.03, rely=0.1, relwidth=0.1)

    namentry = tkinter.Entry(personalbox)
    namentry.place(relx=0.15, rely=0.1, relwidth=0.2)

    sexlabel = tkinter.Label(personalbox, text="Sex", bg='#5F9EA0', fg='#FFFFFF')
    sexlabel.place(relx=0.03, rely=0.2, relwidth=0.1)

    global sexentry
    sexentry = tkinter.Entry(personalbox)
    sexentry.place(relx=0.15, rely=0.2, relwidth=0.2)

    global bloodentry,datentry,placentry
    bloodlabel = tkinter.Label(personalbox, text="Blood Group", bg='#5F9EA0', fg='#FFFFFF')
    bloodlabel.place(relx=0.03, rely=0.3, relwidth=0.1)

    bloodentry = tkinter.Entry(personalbox)
    bloodentry.place(relx=0.15, rely=0.3, relwidth=0.2)

    datelabel = tkinter.Label(personalbox, text="Date of birth", bg='#5F9EA0', fg='#FFFFFF')
    datelabel.place(relx=0.03, rely=0.4, relwidth=0.1)

    datentry = tkinter.Entry(personalbox)
    datentry.place(relx=0.15, rely=0.4, relwidth=0.2)

    placelabel = tkinter.Label(personalbox, text="Place of birth", bg='#5F9EA0', fg='#FFFFFF')
    placelabel.place(relx=0.03, rely=0.5, relwidth=0.1)

    placentry = tkinter.Entry(personalbox)
    placentry.place(relx=0.15, rely=0.5, relwidth=0.2)

    nationlabel = tkinter.Label(personalbox, text="Nationality", bg='#5F9EA0', fg='#FFFFFF')
    nationlabel.place(relx=0.03, rely=0.6, relwidth=0.1)

    global nationentry,townentry,housentry
    nationentry = tkinter.Entry(personalbox)
    nationentry.place(relx=0.15, rely=0.6, relwidth=0.2)

    townlabel = tkinter.Label(personalbox, text="Town", bg='#5F9EA0', fg='#FFFFFF')
    townlabel.place(relx=0.03, rely=0.7, relwidth=0.1)

    townentry = tkinter.Entry(personalbox)
    townentry.place(relx=0.15, rely=0.7, relwidth=0.2)

    houselabel = tkinter.Label(personalbox, text="House Number", bg='#5F9EA0', fg='#FFFFFF')
    houselabel.place(relx=0.03, rely=0.8, relwidth=0.1)

    housentry = tkinter.Entry(personalbox)
    housentry.place(relx=0.15, rely=0.8, relwidth=0.2)

    global mobilentry,telentry,emailentry,departmententry

    mobilelabel = tkinter.Label(personalbox, text="Mobile Number", bg='#5F9EA0', fg='#FFFFFF')
    mobilelabel.place(relx=0.028, rely=0.9, relwidth=0.12)

    mobilentry = tkinter.Entry(personalbox)
    mobilentry.place(relx=0.15, rely=0.9, relwidth=0.2)

    telelabel = tkinter.Label(personalbox, text="Telephone Number", bg='#5F9EA0', fg='#FFFFFF')
    telelabel.place(relx=0.61, rely=0.8, relwidth=0.13)

    telentry = tkinter.Entry(personalbox)
    telentry.place(relx=0.765, rely=0.8, relwidth=0.2)

    emailabel = tkinter.Label(personalbox, text="E-mail", bg='#5F9EA0', fg='#FFFFFF')
    emailabel.place(relx=0.64, rely=0.9, relwidth=0.1)

    emailentry = tkinter.Entry(personalbox)
    emailentry.place(relx=0.765, rely=0.9, relwidth=0.2)

    joboxlabel = tkinter.Label(frame, text='Job detail', bg='#5F9EA0', fg='#FFFFFF')
    joboxlabel.place(relx=0.05, rely=0.7)

    jobox = tkinter.Canvas(frame,height=150, width=800, bd=2, bg='#5F9EA0')# box for job details
    jobox.place(relx=0.05, rely=0.725)

    departmentlabel = tkinter.Label(jobox, text="Department", bg='#5F9EA0', fg='#FFFFFF')
    departmentlabel.place(relx=0.05, rely=0.125, relwidth=0.1)

    departmentview = tkinter.Label(jobox, text=list2[4], bg='#5F9EA0', fg='#FFFFFF')
    departmentview.place(relx=0.15, rely=0.125, relwidth=0.2)

    global joinentry, confirmentry, lastentry, salaryentry, perentry
    joinlabel = tkinter.Label(jobox, text="Date of Joining", bg='#5F9EA0', fg='#FFFFFF')
    joinlabel.place(relx=0.64, rely=0.125, relwidth=0.1)

    joinview = tkinter.Label(jobox, text=list2[5], bg='#5F9EA0', fg='#FFFFFF')
    joinview.place(relx=0.765, rely=0.125, relwidth=0.2)

    confirmlabel = tkinter.Label(jobox, text="Date of confirmation", bg='#5F9EA0', fg='#FFFFFF')
    confirmlabel.place(relx=0.6, rely=0.375, relwidth=0.14)

    confirmview = tkinter.Label(jobox, text=list2[6], bg='#5F9EA0', fg='#FFFFFF')
    confirmview.place(relx=0.765, rely=0.375, relwidth=0.2)

    lastlabel = tkinter.Label(jobox, text="Date of last Increment", bg='#5F9EA0', fg='#FFFFFF')
    lastlabel.place(relx=0.6, rely=0.625, relwidth=0.14)

    lastview = tkinter.Label(jobox, text=list2[7], bg='#5F9EA0', fg='#FFFFFF')
    lastview.place(relx=0.765, rely=0.625, relwidth=0.2)

    salarylabel = tkinter.Label(jobox, text="Salary", bg='#5F9EA0', fg='#FFFFFF')
    salarylabel.place(relx=0.03, rely=0.375, relwidth=0.14)

    salaryview = tkinter.Label(jobox, text=list2[2], bg='#5F9EA0', fg='#FFFFFF')
    salaryview.place(relx=0.15, rely=0.375, relwidth=0.2)

    perlabel = tkinter.Label(jobox, text="Salary per day", bg='#5F9EA0', fg='#FFFFFF')
    perlabel.place(relx=0.03, rely=0.625, relwidth=0.14)

    perview = tkinter.Label(jobox, text=list2[3], bg='#5F9EA0', fg='#FFFFFF' )
    perview.place(relx=0.15, rely=0.625, relwidth=0.2)

    viewbutton = tkinter.Button(frame, text='View', command=displayjob)
    viewbutton.place(relx=0.1, rely=0.955, relwidth=0.1)

    editbutton = tkinter.Button(frame, text='Edit', command=update)
    editbutton.place(relx=0.225, rely=0.955, relwidth=0.1)

    cancelbutton = tkinter.Button(frame, text='Cancel', command = cancel)
    cancelbutton.place(relx=0.35, rely=0.955, relwidth=0.1)

    canvas.pack()

    profwindow.mainloop()
    return profwindow

def jobEdit():
    global win3_count
    win3_count+=1
    if win3_count >1:
        messagebox.showinfo("warning","Window is already opened")
        return
    def cancel():
        global win_count
        win3_count = 0
        return jobedwindow.destroy()
    excel=('HRM_Excel.xlsx')
    wb=xlrd.open_workbook(excel)
    sheet=wb.sheet_by_name("Employee Personal Information")

    list1=[]                #list of id
    for i in range(2,sheet.nrows):
        list1.append(sheet.cell_value(i,0))

    global jobedwindow,gradentry,designationentry
    jobedwindow = tkinter.Tk()
    jobedwindow.title("Update employee's Job Data")
    canvas = tkinter.Canvas(jobedwindow, height=700, width=900) 

    frame = tkinter.Frame(jobedwindow, bg='#5F9EA0')
    frame.place(relwidth=1, relheight=1)

    boxlabel = tkinter.Label(frame, text='Main Profile', bg='#5F9EA0', fg='#FFFFFF')
    boxlabel.place(relx=0.05, rely=0.025)

    mainbox = tkinter.Canvas(frame, height=100, width=800, bd=2, bg='#5F9EA0')# the box of main profile
    mainbox.place(relx=0.05, rely=0.05)

    global idcombo,namentry,typentry
    idlabel = tkinter.Label(mainbox, text="Employee ID", bg='#5F9EA0', fg='#FFFFFF')
    idlabel.place(relx=0.03, rely=0.2, relwidth=0.14)


    idcombo=Combobox(mainbox,values=list1)
    idcombo.place(relx=0.19, rely=0.2, relwidth=0.2)

    list2=["-","-","-","-","-","-","-","-","-","-","-","-"]

    def displayinfo():
        list2.clear()
        for i in range(1,sheet.nrows):          
            if idcombo.get()==sheet.cell_value(i,0):
                
                list2.append(sheet.cell_value(i,1))
                designationview = tkinter.Label(mainbox, text=list2[0], bg='#5F9EA0', fg='#FFFFFF')
                designationview.place(relx=0.15, rely=0.6, relwidth=0.2)

                list2.append(sheet.cell_value(i,2))
                empnameview = tkinter.Label(mainbox, text=list2[1], bg='#5F9EA0', fg='#FFFFFF')
                empnameview.place(relx=0.75, rely=0.2, relwidth=0.2)
                nameview = tkinter.Label(personalbox, text=list2[1], bg='#5F9EA0', fg='#FFFFFF')
                nameview.place(relx=0.15, rely=0.1, relwidth=0.2)

                list2.append(sheet.cell_value(i,3))
                sexview = tkinter.Label(personalbox, text=list2[2], bg='#5F9EA0', fg='#FFFFFF')
                sexview.place(relx=0.15, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,4))
                bloodview = tkinter.Label(personalbox, text=list2[3], bg='#5F9EA0', fg='#FFFFFF')
                bloodview.place(relx=0.15, rely=0.3, relwidth=0.2)

                list2.append(sheet.cell_value(i,5))
                dateview = tkinter.Label(personalbox, text=list2[4], bg='#5F9EA0', fg='#FFFFFF')
                dateview.place(relx=0.15, rely=0.4, relwidth=0.2)

                list2.append(sheet.cell_value(i,6))
                placeview = tkinter.Label(personalbox, text=list2[5], bg='#5F9EA0', fg='#FFFFFF')
                placeview.place(relx=0.15, rely=0.5, relwidth=0.2)

                list2.append(sheet.cell_value(i,7))
                nationview = tkinter.Label(personalbox, text=list2[6], bg='#5F9EA0', fg='#FFFFFF')
                nationview.place(relx=0.1, rely=0.6, relwidth=0.3)

                list2.append(sheet.cell_value(i,8))
                townview = tkinter.Label(personalbox, text=list2[7], bg='#5F9EA0', fg='#FFFFFF')
                townview.place(relx=0.15, rely=0.7, relwidth=0.2)

                list2.append(sheet.cell_value(i,9))
                houseview = tkinter.Label(personalbox, text=list2[8], bg='#5F9EA0', fg='#FFFFFF')
                houseview.place(relx=0.15, rely=0.8, relwidth=0.2)

                list2.append(sheet.cell_value(i,10))
                mobileview = tkinter.Label(personalbox, text=list2[9], bg='#5F9EA0', fg='#FFFFFF')
                mobileview.place(relx=0.15, rely=0.9, relwidth=0.2)

                list2.append(sheet.cell_value(i,11))
                teleview = tkinter.Label(personalbox, text=list2[10], bg='#5F9EA0', fg='#FFFFFF')
                teleview.place(relx=0.765, rely=0.8, relwidth=0.2)

                list2.append(sheet.cell_value(i,12))
                emailview = tkinter.Label(personalbox, text=list2[11], bg='#5F9EA0', fg='#FFFFFF')
                emailview.place(relx=0.765, rely=0.9, relwidth=0.2)


                list2.append(sheet.cell_value(i,12))


    designationlabel = tkinter.Label(mainbox, text="Position", bg='#5F9EA0', fg='#FFFFFF')
    designationlabel.place(relx=0.03, rely=0.6, relwidth=0.1)

    designationview = tkinter.Label(mainbox,text=list2[0], bg='#5F9EA0')
    designationview.place(relx=0.15, rely=0.6, relwidth=0.2)

    empnamelabel = tkinter.Label(mainbox, text="employee name", bg='#5F9EA0', fg='#FFFFFF')
    empnamelabel.place(relx=0.6, rely=0.2, relwidth=0.14)

    empnameview = tkinter.Label(mainbox,text=list2[1], bg='#5F9EA0')
    empnameview.place(relx=0.75, rely=0.2, relwidth=0.2)

    personalboxlabel = tkinter.Label(frame, text='Personal Profile', bg='#5F9EA0', fg='#FFFFFF')
    personalboxlabel.place(relx=0.05, rely=0.225)

    personalbox = tkinter.Canvas(frame,height=300, width=800, bd=2, bg='#5F9EA0')# the box for personal profile
    personalbox.place(relx=0.05, rely=0.25)

    namelabel = tkinter.Label(personalbox, text="Name", bg='#5F9EA0', fg='#FFFFFF')
    namelabel.place(relx=0.03, rely=0.1, relwidth=0.1)

    nameview = tkinter.Label(personalbox,text=list2[1], bg='#5F9EA0')
    nameview.place(relx=0.15, rely=0.1, relwidth=0.2)

    sexlabel = tkinter.Label(personalbox, text="Sex", bg='#5F9EA0', fg='#FFFFFF')
    sexlabel.place(relx=0.03, rely=0.2, relwidth=0.1)

    sexview = tkinter.Label(personalbox,text=list2[2], bg='#5F9EA0')
    sexview.place(relx=0.15, rely=0.3, relwidth=0.2)

    #global bloodentry,datentry,placentry
    bloodlabel = tkinter.Label(personalbox, text="Blood Group", bg='#5F9EA0', fg='#FFFFFF')
    bloodlabel.place(relx=0.03, rely=0.3, relwidth=0.1)

    bloodview = tkinter.Label(personalbox,text=list2[3], bg='#5F9EA0')
    bloodview.place(relx=0.15, rely=0.3, relwidth=0.2)

    datelabel = tkinter.Label(personalbox, text="Date of birth", bg='#5F9EA0', fg='#FFFFFF')
    datelabel.place(relx=0.03, rely=0.4, relwidth=0.1)

    dateview = tkinter.Label(personalbox,text=list2[4], bg='#5F9EA0')
    dateview.place(relx=0.15, rely=0.4, relwidth=0.2)

    placelabel = tkinter.Label(personalbox, text="Place of birth", bg='#5F9EA0', fg='#FFFFFF')
    placelabel.place(relx=0.03, rely=0.5, relwidth=0.1)

    placeview = tkinter.Label(personalbox,text=list2[5], bg='#5F9EA0')
    placeview.place(relx=0.15, rely=0.5, relwidth=0.2)

    nationlabel = tkinter.Label(personalbox, text="Nationality", bg='#5F9EA0', fg='#FFFFFF')
    nationlabel.place(relx=0.03, rely=0.6, relwidth=0.1)

    global nationentry,townentry,housentry
    nationview = tkinter.Label(personalbox,text=list2[6], bg='#5F9EA0')
    nationview.place(relx=0.15, rely=0.6, relwidth=0.3)

    townlabel = tkinter.Label(personalbox, text="Town", bg='#5F9EA0', fg='#FFFFFF')
    townlabel.place(relx=0.03, rely=0.7, relwidth=0.1)

    townview = tkinter.Label(personalbox,text=list2[7], bg='#5F9EA0')
    townview.place(relx=0.15, rely=0.7, relwidth=0.2)

    houselabel = tkinter.Label(personalbox, text="House Number", bg='#5F9EA0', fg='#FFFFFF')
    houselabel.place(relx=0.03, rely=0.8, relwidth=0.1)

    housview = tkinter.Label(personalbox,text=list2[8], bg='#5F9EA0')
    housview.place(relx=0.15, rely=0.8, relwidth=0.2)

    global mobilentry,telentry,emailentry,departmententry
    mobilelabel = tkinter.Label(personalbox, text="Mobile Number", bg='#5F9EA0', fg='#FFFFFF')
    mobilelabel.place(relx=0.028, rely=0.9, relwidth=0.12)

    mobileview = tkinter.Label(personalbox,text=list2[9], bg='#5F9EA0')
    mobileview.place(relx=0.15, rely=0.9, relwidth=0.2)

    telelabel = tkinter.Label(personalbox, text="Telephone Number", bg='#5F9EA0', fg='#FFFFFF')
    telelabel.place(relx=0.61, rely=0.8, relwidth=0.13)

    teleview = tkinter.Label(personalbox,text=list2[10], bg='#5F9EA0')
    teleview.place(relx=0.765, rely=0.8, relwidth=0.2)

    emailabel = tkinter.Label(personalbox, text="E-mail", bg='#5F9EA0', fg='#FFFFFF')
    emailabel.place(relx=0.64, rely=0.9, relwidth=0.1)

    emailview = tkinter.Label(personalbox,text=list2[11], bg='#5F9EA0')
    emailview.place(relx=0.765, rely=0.9, relwidth=0.2)
    
    global joinentry, confirmentry, lastentry, salaryentry, perentry
    joboxlabel = tkinter.Label(frame, text='Job detail', bg='#5F9EA0', fg='#FFFFFF')
    joboxlabel.place(relx=0.05, rely=0.7)

    jobox = tkinter.Canvas(frame,height=150, width=800, bd=2, bg='#5F9EA0')# box for job details
    jobox.place(relx=0.05, rely=0.725)

    departmentlabel = tkinter.Label(jobox, text="Department", bg='#5F9EA0', fg='#FFFFFF')
    departmentlabel.place(relx=0.05, rely=0.125, relwidth=0.1)

    departmententry = tkinter.Entry(jobox)
    departmententry.place(relx=0.15, rely=0.125, relwidth=0.2)

    joinlabel = tkinter.Label(jobox, text="Date of Joining", bg='#5F9EA0', fg='#FFFFFF')
    joinlabel.place(relx=0.64, rely=0.125, relwidth=0.1)

    joinentry = tkinter.Entry(jobox)
    joinentry.place(relx=0.765, rely=0.125, relwidth=0.2)

    confirmlabel = tkinter.Label(jobox, text="Date of confirmation", bg='#5F9EA0', fg='#FFFFFF')
    confirmlabel.place(relx=0.6, rely=0.375, relwidth=0.14)

    confirmentry = tkinter.Entry(jobox)
    confirmentry.place(relx=0.765, rely=0.375, relwidth=0.2)

    lastlabel = tkinter.Label(jobox, text="Date of last Increment", bg='#5F9EA0', fg='#FFFFFF')
    lastlabel.place(relx=0.6, rely=0.625, relwidth=0.14)

    lastentry = tkinter.Entry(jobox)
    lastentry.place(relx=0.765, rely=0.625, relwidth=0.2)

    salarylabel = tkinter.Label(jobox, text="Salary", bg='#5F9EA0', fg='#FFFFFF')
    salarylabel.place(relx=0.03, rely=0.375, relwidth=0.14)

    salaryentry = tkinter.Entry(jobox)
    salaryentry.place(relx=0.15, rely=0.375, relwidth=0.2)

    perlabel = tkinter.Label(jobox, text="Salary per day", bg='#5F9EA0', fg='#FFFFFF')
    perlabel.place(relx=0.03, rely=0.625, relwidth=0.14)

    perentry = tkinter.Entry(jobox)
    perentry.place(relx=0.15, rely=0.625, relwidth=0.2)

    viewbutton = tkinter.Button(frame, text='View', command=displayinfo)
    viewbutton.place(relx=0.1, rely=0.955, relwidth=0.1)

    editbutton = tkinter.Button(frame, text="Edit", command=update2)
    editbutton.place(relx=0.225, rely=0.955, relwidth=0.1)

    cancelbutton = tkinter.Button(frame, text='Cancel', command = cancel)
    cancelbutton.place(relx=0.35, rely=0.955, relwidth=0.1)

    canvas.pack()

    jobedwindow.mainloop()
    return jobedwindow


def getinput():
    global designation,Name,Sex,Bld_grp,DOB,POB,Nat,town,Hu_No,Mb_No,Tele_no,Email

    #Birthdate - month/day/year
    global birthdate, month, day, year
    month, day, year = monthcombo.get(), datecombo.get(), yearcombo.get()
    birthdate = str(month) + "/" + str(day) + "/" + str(year)
    
    designation,Name,Sex,Bld_grp = designationentry.get(),namentry.get(),sexcombo.get(),bloodcombo.get()
    DOB,POB,Nat,town = birthdate,placentry.get(),nationentry.get(),townentry.get()
    Hu_No,Mb_No,Tele_no,Email = housentry.get(),mobilentry.get(),telentry.get(),emailentry.get()

    
    global department,salary,salaryperday,datejoined,dateconfirmed,dateincrement
    
    department,salary,salaryperday = departmententry.get(),salaryentry.get(),perentry.get()
    datejoined,dateconfirmed,dateincrement = joinentry.get(), confirmentry.get(), lastentry.get()

    if designation == '' or Name == '' or Sex == '' or Bld_grp == '' or month == '' or day == '' or year == '' \
    or POB == '' or Nat == '' or town == '' or Hu_No == '' or Mb_No == '' or Tele_no == '' or Email == '' or \
    department == '' or salary == '' or salaryperday == '' or datejoined == '' or dateconfirmed == ''or dateincrement == '':
        messagebox.showinfo("Incomplete","Please fill out all entries!")

    else:
        messagebox.showinfo("Saved","Succesfully Saved!")
        addwindow.destroy()
    
    excel()

def excel():
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Employee Personal Information"]
    data = [[idgen,designation,Name,Sex,Bld_grp,DOB,POB,Nat,town,Hu_No,Mb_No,Tele_no,Email]]
    
    for i in data:
        if designation != '' and Name != '' and Sex != '' and Bld_grp != '' and month != '' and day != '' and year != '' \
        or POB != '' and Nat != '' and town != '' and Hu_No != '' and Mb_No != '' and Tele_no != '' and Email != '' and \
        department != '' and salary != '' and salaryperday != '' and datejoined != '' and dateconfirmed != ''and dateincrement != '':
            ws.append(i)
            wb.save("HRM_Excel.xlsx")
    
    ws2 = wb["Employee Job Details"]
    data2 = [[idgen,designation,Name,department,salary,salaryperday,datejoined,dateconfirmed,dateincrement]]
    
    for i in data2:
        if designation != '' and Name != '' and Sex != '' and Bld_grp != '' and month != '' and day != '' and year != '' \
        or POB != '' and Nat != '' and town != '' and Hu_No != '' and Mb_No != '' and Tele_no != '' and Email != '' and \
        department != '' and salary != '' and salaryperday != '' and datejoined != '' and dateconfirmed != ''and dateincrement != '':
                ws2.append(i)
                wb.save("HRM_Excel.xlsx")

    data3 = [[idgen]]
    ws3 = wb["Charts"]
    for i in data3:
        ws3.append(i)
        wb.save("HRM_Excel.xlsx")

    data4 = [[idgen,designation,Name,Sex,Bld_grp,DOB,POB,Nat,town,Hu_No,Mb_No,Tele_no,Email,department,salary,salaryperday\
              ,datejoined,dateconfirmed,dateincrement]]
    
    ws4 = wb["Merged"]
    for i in data4:
        ws4.append(i)
        wb.save("HRM_Excel.xlsx")
    
def search(x):
    global b
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Employee Personal Information"]
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == idcombo.get():
                for cell2 in row:
                    count+=1
                    if count == x:
                        b = cell2.coordinate
                        break
                break
    return b
def update(): 
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Employee Personal Information"]
    cellname = ws[search(3)] 
    cellname.value = namentry.get()
    sexname = ws[search(4)]
    sexname.value = sexentry.get()
    blood_g = ws[search(5)]
    blood_g.value = bloodentry.get()
    DOB = ws[search(6)]
    DOB.value = datentry.get()
    POB = ws[search(7)]
    POB.value = placentry.get()
    nat = ws[search(8)]
    nat.value = nationentry.get()
    Tow = ws[search(9)]
    Tow.value = townentry.get()
    hu_nu = ws[search(10)]
    hu_nu.value = housentry.get()
    mo_nu = ws[search(11)]
    mo_nu.value = mobilentry.get()
    tele = ws[search(12)]
    tele.value = telentry.get()
    e_mail = ws[search(13)]
    e_mail.value = emailentry.get()
    
    wb.save("HRM_Excel.xlsx")
    messagebox.showinfo('success','Successfully Updated!')
def search2(x):
    global a
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Employee Job Details"]
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == idcombo.get():
                for cell2 in row:
                    count+=1
                    if count == x:
                        a = cell2.coordinate
                        break
                break
    return a

def update2():
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Employee Job Details"]
    celldep = ws[search2(4)]
    celldep.value = departmententry.get()
    salar = ws[search2(5)]
    salar.value = salaryentry.get()
    daysar = ws[search2(6)]
    daysar.value = perentry.get()
    DOJ = ws[search2(7)]
    DOJ.value = joinentry.get()
    DOC = ws[search2(8)]
    DOC.value = confirmentry.get()
    DOI = ws[search2(9)]
    DOI.value = lastentry.get()
    wb.save("HRM_Excel.xlsx")
    messagebox.showinfo('success',"Successfully Updated")
    
def leaveApprove():
    excel=('HRM_Excel.xlsx')
    wb=xlrd.open_workbook(excel)
    sheet=wb.sheet_by_name("Leaves")

    list1=[]                
    for i in range(2,sheet.nrows):
        list1.append(sheet.cell_value(i,0))

    window = tkinter.Tk()
    window.title("leave Application")
    window.resizable(False, False)
    canvas = tkinter.Canvas(window, height=700, width=900) 

    frame = tkinter.Frame(window, bg='#5F9EA0')
    frame.place(relwidth=1, relheight=1)

    boxlabel = tkinter.Label(frame, text='Main Profile', bg='#5F9EA0', fg='#FFFFFF')
    boxlabel.place(relx=0.05, rely=0.025)

    mainbox = tkinter.Canvas(frame, height=100, width=800, bd=2, bg='#5F9EA0')# the box of main profile
    mainbox.place(relx=0.05, rely=0.05)

    idlabel = tkinter.Label(mainbox, text="Employee ID", bg='#5F9EA0', fg='#FFFFFF')
    idlabel.place(relx=0.03, rely=0.2, relwidth=0.14)
    global combo
    combo=Combobox(mainbox,values=list1,width=20)
    combo.place(relx=0.19, rely=0.2, relwidth=0.2)

    list2=["-","-","-","-","-","-","-","-","-"]
    def leave():
        list2.clear()
        for i in range(2,sheet.nrows):          
            if combo.get()==sheet.cell_value(i,0):

                list2.append(sheet.cell_value(i,1))
                designationlabel = tkinter.Label(mainbox, text=list2[0],  bg='#5F9EA0', fg='#FFFFFF')
                designationlabel.place(relx=0.15, rely=0.6, relwidth=0.2)

                list2.append(sheet.cell_value(i,2))
                empnamelabel = tkinter.Label(mainbox,text=list2[1],  bg='#5F9EA0', fg='#FFFFFF')
                empnamelabel.place(relx=0.75, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,3))
                monthlabel = tkinter.Label(leavebox,text=list2[2],  bg='#5F9EA0', fg='#FFFFFF')
                monthlabel.place(relx=0.15, rely=0.1, relwidth=0.2)

                list2.append(sheet.cell_value(i,4))
                startlabel = tkinter.Label(leavebox, text=list2[3],  bg='#5F9EA0', fg='#FFFFFF')
                startlabel.place(relx=0.15, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,5))
                applylabel = tkinter.Label(leavebox, text=list2[4],  bg='#5F9EA0', fg='#FFFFFF')
                applylabel.place(relx=0.15, rely=0.3, relwidth=0.2)

                list2.append(sheet.cell_value(i,6))
                endlabel = tkinter.Label(leavebox, text=list2[5],  bg='#5F9EA0', fg='#FFFFFF')
                endlabel.place(relx=0.765, rely=0.1, relwidth=0.2)

                list2.append(sheet.cell_value(i,7))
                periodlabel = tkinter.Label(leavebox, text=list2[6],  bg='#5F9EA0', fg='#FFFFFF')
                periodlabel.place(relx=0.765, rely=0.2, relwidth=0.2)
                
                list2.append(sheet.cell_value(i,8))
                
            

    designationlabel = tkinter.Label(mainbox, text="Position", bg='#5F9EA0', fg='#FFFFFF')
    designationlabel.place(relx=0.03, rely=0.6, relwidth=0.1)

    designationlabel = tkinter.Label(mainbox,text=list2[0],  bg='#5F9EA0', fg='#FFFFFF')
    designationlabel.place(relx=0.15, rely=0.6, relwidth=0.2)

    empnamelabel = tkinter.Label(mainbox, text="Employee Name", bg='#5F9EA0', fg='#FFFFFF')
    empnamelabel.place(relx=0.6, rely=0.2, relwidth=0.14)

    empnamelabel = tkinter.Label(mainbox, text=list2[1],  bg='#5F9EA0', fg='#FFFFFF')
    empnamelabel.place(relx=0.75, rely=0.2, relwidth=0.2)

    leaveboxlabel = tkinter.Label(frame, text='Leave application', bg='#5F9EA0', fg='#FFFFFF')
    leaveboxlabel.place(relx=0.05, rely=0.225)

    leavebox = tkinter.Canvas(frame,height=450, width=800, bd=2, bg='#5F9EA0')# the box for leave application form
    leavebox.place(relx=0.05, rely=0.25)

    monthlabel = tkinter.Label(leavebox, text="Month", bg='#5F9EA0', fg='#FFFFFF')
    monthlabel.place(relx=0.03, rely=0.1, relwidth=0.1)

    monthlabel = tkinter.Label(leavebox, text=list2[2],  bg='#5F9EA0', fg='#FFFFFF')
    monthlabel.place(relx=0.15, rely=0.1, relwidth=0.2)

    startlabel = tkinter.Label(leavebox, text="Start date", bg='#5F9EA0', fg='#FFFFFF')
    startlabel.place(relx=0.03, rely=0.2, relwidth=0.1)

    startlabel = tkinter.Label(leavebox, text=list2[3],  bg='#5F9EA0', fg='#FFFFFF')
    startlabel.place(relx=0.15, rely=0.2, relwidth=0.2)

    endlabel = tkinter.Label(leavebox, text="End date", bg='#5F9EA0', fg='#FFFFFF')
    endlabel.place(relx=0.64, rely=0.1, relwidth=0.1)

    endlabel = tkinter.Label(leavebox, text=list2[5],  bg='#5F9EA0', fg='#FFFFFF')
    endlabel.place(relx=0.765, rely=0.1, relwidth=0.2)

    periodlabel = tkinter.Label(leavebox, text="Leave period", bg='#5F9EA0', fg='#FFFFFF')
    periodlabel.place(relx=0.64, rely=0.2, relwidth=0.1)

    periodlabel = tkinter.Label(leavebox, text=list2[6],  bg='#5F9EA0', fg='#FFFFFF')
    periodlabel.place(relx=0.765, rely=0.2, relwidth=0.2)

    applylabel = tkinter.Label(leavebox, text="Reason For Leave", bg='#5F9EA0', fg='#FFFFFF')
    applylabel.place(relx=0.02, rely=0.3, relwidth=0.13)

    applylabel = tkinter.Label(leavebox,text=list2[4],  bg='#5F9EA0', fg='#FFFFFF')
    applylabel.place(relx=0.15, rely=0.3, relwidth=0.1)
    
    viewbutton = tkinter.Button(frame, text='View', command=leave)
    viewbutton.place(relx=0.2, rely=0.9125, relwidth=0.1)

    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Leaves"]
    def search3(x):
        global c
        wb = load_workbook('HRM_Excel.xlsx')
        ws = wb["Leaves"]
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == combo.get():
                    for cell2 in row:
                        count+=1
                        if count == x:
                            c = cell2.coordinate
                            print (c)
                            break
                    break
        return c
    def approv():
        global cellstatus
        cellstatus = ws[search3(10)]
        cellstatus.value = "Appproved"
        wb.save("HRM_Excel.xlsx")
        return cellstatus.value
    def denied():
        global cellstatus
        cellstatus = ws[search3(10)]
        cellstatus.value = "Denied"
        wb.save("HRM_Excel.xlsx")
        window.destroy()
        
    approvebutton = tkinter.Button(frame, text='Approve', command=approv)
    approvebutton.place(relx=0.325, rely=0.9125, relwidth=0.1)

    denybutton = tkinter.Button(frame, text='Deny', command = denied)
    denybutton.place(relx=0.45, rely=0.9125, relwidth=0.1)

    canvas.pack()

    window.mainloop()
    return window
    
def atendance():
    excel=('HRM_Excel.xlsx')
    wb=xlrd.open_workbook(excel)
    sheet=wb.sheet_by_name("Attendance")


    list1=[]                
    for i in range(2,sheet.nrows):
        list1.append(sheet.cell_value(i,0))

    global attendwindow
    attendwindow = tkinter.Tk()
    attendwindow.resizable(False, False)
    attendwindow.title("Attendance Detail")
    canvas = tkinter.Canvas(attendwindow, height=700, width=900) 

    frame = tkinter.Frame(attendwindow, bg='#5F9EA0')
    frame.place(relwidth=1, relheight=1)

    boxlabel = tkinter.Label(frame, text='Main Profile', bg='#5F9EA0', fg='#FFFFFF')
    boxlabel.place(relx=0.05, rely=0.025)

    mainbox = tkinter.Canvas(frame, height=100, width=800, bd=2, bg='#5F9EA0') #the box of main profile
    mainbox.place(relx=0.05, rely=0.05)

    global identry, gradentry, designationentry, typentry 
    
    idlabel = tkinter.Label(mainbox, text="Employee ID", bg='#5F9EA0', fg='#FFFFFF')
    idlabel.place(relx=0.03, rely=0.2, relwidth=0.14)

    global combo
    combo=Combobox(mainbox,values=list1,width=20)
    combo.place(relx=0.19, rely=0.2, relwidth=0.2)

    list2=["-","-","-","-","-","-","-","-","-"]
    def attend():
        list2.clear()
        for i in range(2,sheet.nrows):          
            if combo.get()==sheet.cell_value(i,0):
                list2.append(sheet.cell_value(i,1))
                designationlabel = tkinter.Label(mainbox, text=list2[0],  bg='#5F9EA0', fg='#FFFFFF')
                designationlabel.place(relx=0.19, rely=0.6, relwidth=0.2)

                list2.append(sheet.cell_value(i,2))
                empnamelabel = tkinter.Label(mainbox, text=list2[1],  bg='#5F9EA0', fg='#FFFFFF')
                empnamelabel.place(relx=0.765, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,3))
                workentry = tkinter.Label(atendancebox, text=list2[2],  bg='#5F9EA0', fg='#FFFFFF')
                workentry.place(relx=0.2, rely=0.1, relwidth=0.2)

                list2.append(sheet.cell_value(i,4))
                presententry = tkinter.Label(atendancebox, text=list2[3],  bg='#5F9EA0', fg='#FFFFFF')
                presententry.place(relx=0.2, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,5))        
                monthentry = tkinter.Label(atendancebox, text=list2[4],  bg='#5F9EA0', fg='#FFFFFF')
                monthentry.place(relx=0.765, rely=0.1, relwidth=0.2)

                list2.append(sheet.cell_value(i,6))
                yearentry = tkinter.Label(atendancebox, text=list2[5],  bg='#5F9EA0', fg='#FFFFFF')
                yearentry.place(relx=0.765, rely=0.2, relwidth=0.2)

                list2.append(sheet.cell_value(i,7))
                reportentry.place(relx=0.03, rely=0.34, relwidth=0.9, relheight=0.6)
                list2.append(sheet.cell_value(i,8))

    designationlabel = tkinter.Label(mainbox, text="Position", bg='#5F9EA0', fg='#FFFFFF')
    designationlabel.place(relx=0.03, rely=0.6, relwidth=0.14)

    designationlabel = tkinter.Label(mainbox)
    designationlabel.place(relx=0.19, rely=0.6, relwidth=0.2)

    empnamelabel = tkinter.Label(mainbox, text="Employee", bg='#5F9EA0', fg='#FFFFFF')
    empnamelabel.place(relx=0.64, rely=0.2, relwidth=0.1)

    empnamelabel = tkinter.Label(mainbox)
    empnamelabel.place(relx=0.765, rely=0.2, relwidth=0.2)

    global workentry, presententry, monthentry, yearentry, reportentry
    
    atendanceboxlabel = tkinter.Label(frame, text='Attendance detail', bg='#5F9EA0', fg='#FFFFFF')
    atendanceboxlabel.place(relx=0.05, rely=0.225)

    atendancebox = tkinter.Canvas(frame,height=470, width=800, bd=2, bg='#5F9EA0')# the box for atendance
    atendancebox.place(relx=0.05, rely=0.25)

    worklabel = tkinter.Label(atendancebox, text="Total working days", bg='#5F9EA0', fg='#FFFFFF')
    worklabel.place(relx=0.03, rely=0.1, relwidth=0.15)

    workentry = tkinter.Label(atendancebox)
    workentry.place(relx=0.2, rely=0.1, relwidth=0.2)

    presentlabel = tkinter.Label(atendancebox, text="Total present days", bg='#5F9EA0', fg='#FFFFFF')
    presentlabel.place(relx=0.03, rely=0.2, relwidth=0.15)

    presententry = tkinter.Label(atendancebox)
    presententry.place(relx=0.2, rely=0.2, relwidth=0.2)

    monthlabel = tkinter.Label(atendancebox, text="Month", bg='#5F9EA0', fg='#FFFFFF')
    monthlabel.place(relx=0.64, rely=0.1, relwidth=0.1)

    monthentry = tkinter.Label(atendancebox)
    monthentry.place(relx=0.765, rely=0.1, relwidth=0.2)

    yearlabel = tkinter.Label(atendancebox, text="Year", bg='#5F9EA0', fg='#FFFFFF')
    yearlabel.place(relx=0.64, rely=0.2, relwidth=0.1)

    yearentry = tkinter.Label(atendancebox)
    yearentry.place(relx=0.765, rely=0.2, relwidth=0.2)

    reportlabel = tkinter.Label(atendancebox, text="Monthly atendance Report", bg='#5F9EA0', fg='#FFFFFF')
    reportlabel.place(relx=0.03, rely=0.3, relwidth=0.2)

    reportentry = tkinter.Entry(atendancebox)
    reportentry.place(relx=0.03, rely=0.34, relwidth=0.9, relheight=0.6)

    applybutton = tkinter.Button(frame, text='View', command=attend)
    applybutton.place(relx=0.1, rely=0.935, relwidth=0.1)

    def search4(x):
        global d
        wb = load_workbook('HRM_Excel.xlsx')
        ws = wb["Attendance"]
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == combo.get():
                    for cell2 in row:
                        count+=1
                        if count == x:
                            d = cell2.coordinate
                            break
                    break
        return d
    def sav():
        global cellstatus
        wb = load_workbook('HRM_Excel.xlsx')
        ws = wb["Attendance"]
        cellstatus = ws[search4(8)]
        cellstatus.value = "checked"
        wb.save("HRM_Excel.xlsx")
        messagebox.showinfo('Success',"Succesfully Saved")
        return cellstatus.value
    applybutton = tkinter.Button(frame, text='Save', command = sav)
    applybutton.place(relx=0.225, rely=0.935, relwidth=0.1)

    applybutton = tkinter.Button(frame, text='Cancel', command = attendwindow.destroy )
    applybutton.place(relx=0.35, rely=0.935, relwidth=0.1)

    canvas.pack()

    attendwindow.mainloop()
    return attendwindow

def getinput5():
    global ID,grade,designation,Type,work,present,month,year,report
    
    ID,grade,designation,Type,work = idgen, gradentry.get(), designationentry.get(),typentry.get(),workentry.get()
    present,month,year,report = presententry.get(),monthentry.get(),yearentry.get(),reportentry.get()

    excel5()
    
def excel5():
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Attendance"]
    data = [[ID,grade,designation,Type,work,present,month,year,report]]
    for i in data:
        ws.append(i)
    wb.save("HRM_Excel.xlsx")
    
def announcement():
    excel=('HRM_Excel.xlsx')
    wb=xlrd.open_workbook(excel)
    sheet=wb.sheet_by_name("Employee Personal Information")


    list1=[] 
    for i in range(2,sheet.nrows):
        list1.append(sheet.cell_value(i,2))

    global anouncewindow

    anouncewindow = tkinter.Tk()
    anouncewindow.title("Send Announcement")
    anouncewindow.resizable(False, False)
    canvas = tkinter.Canvas(anouncewindow, height=700, width=900)

    frame = tkinter.Frame(anouncewindow, bg='#5F9EA0')
    frame.place(relwidth=1, relheight=1)

    boxlabel = tkinter.Label(frame, text='Announcement', bg='#5F9EA0', fg='#FFFFFF')
    boxlabel.place(relx=0.05, rely=0.025)

    mainbox = tkinter.Canvas(frame, height=600, width=800, bd=2, bg='#5F9EA0')# the box of main profile
    mainbox.place(relx=0.05, rely=0.05)

    global  annamecombo, subjectentry, messagentry
    tolabel = tkinter.Label(mainbox, text="To", bg='#5F9EA0', fg='#FFFFFF')
    tolabel.place(relx=0.03, rely=0.1, relwidth=0.1)

    annamecombo = Combobox(mainbox,values=list1,width=20)
    annamecombo.place(relx=0.15, rely=0.1, relwidth=0.5)

    subjectlabel = tkinter.Label(mainbox, text="Subject", bg='#5F9EA0', fg='#FFFFFF')
    subjectlabel.place(relx=0.03, rely=0.2, relwidth=0.1)

    subjectentry = tkinter.Entry(mainbox)
    subjectentry.place(relx=0.15, rely=0.2, relwidth=0.5)

    messagelabel = tkinter.Label(mainbox, text="Message", bg='#5F9EA0', fg='#FFFFFF')
    messagelabel.place(relx=0.03, rely=0.3, relwidth=0.1)

    messagentry = tkinter.Entry(mainbox)
    messagentry.place(relx=0.15, rely=0.3, relwidth=0.5, relheight=0.5)

    savebutton = tkinter.Button(frame, text='Send', command = getinput4)
    savebutton.place(relx=0.35, rely=0.8225, relwidth=0.1)
    
    canvas.pack()

    anouncewindow.mainloop()
    return anouncewindow

def getinput4():
    global to, subject, message
    to, subject, message = annamecombo.get(), subjectentry.get(), messagentry.get()

    excel4()
    
    
def excel4():
    wb = load_workbook('HRM_Excel.xlsx')
    ws = wb["Announcement"]
    data = [[to, subject, message]]

    for i in data:
        ws.append(i)

    if to == '' or subject == '' or message == '':
        messagebox.showinfo('Incomplete','Please fill out all entries.')
        
    if to != '' and subject != '' and message != '':
        messagebox.showinfo('Sent','Announcement has been sent')
        anouncewindow.destroy()

    wb.save("HRM_Excel.xlsx")

    
            
def Logout():
    def LG():
        Home_window.destroy()
        win.destroy()
        
    global win
    win=tkinter.Tk()
    win.geometry('150x150')
    win.title('Logout')
    tkinter.Label(win, text="Are you sure?").pack()
    tkinter.Button(win, text="Ok", command= LG).place(x = '50', y = '30')
    tkinter.Button(win, text="Cancel", command= win.destroy).place(x = '50', y = '70')
        
    win.mainloop()
    return win

def mainscreen():
    global win_count,win1_count,win3_count
    win_count = 0
    win1_count = 0
    win3_count = 0
    global Home_window
    Home_window = tkinter.Toplevel()#creats an empty window
    #Home_window = Tk()

    Home_window.title("Admin Home page")
    Home_window.geometry("1250x650")
    Home_window.resizable(False, False)
    menubar = tkinter.Menu(Home_window)
    Home_window.config(menu=menubar)
    Employee = tkinter.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Employee", menu=Employee)
    Employee.add_command(label="Add New Employee", command=profileNew)
    Employee.add_cascade(label="Update Employee Personal Data", command=profileEdit)
    Employee.add_cascade(label="Update Employee Job Data", command=jobEdit)

    def lock():
        passw = tkinter.simpledialog.askinteger("Data password", "Enter new password")
        if passw == None:
            return
        wb = load_workbook('HRM_Excel.xlsx')
        #ws = wb['Employee Personal Information']
        list1 = wb.sheetnames
        for i in list1:
            ws = wb[i]
            ws.protection.password = str(passw)
        tkinter.messagebox.showinfo("confirmation", "password changed!")
        #print(list1)
        wb.save("HRM_Excel.xlsx")
    leavemenu = tkinter.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Leaves", menu=leavemenu)
    leavemenu.add_command(label="Leave Application", command=leaveApprove)


    AttendanceMenu = tkinter.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Attendance", menu=AttendanceMenu)
    AttendanceMenu.add_command(label=" Attendace ", command=atendance)

    Announcementmenu = tkinter.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Announcement", menu=Announcementmenu)
    Announcementmenu.add_command(label="Send Announcement", command=announcement)

    Logoutmenu = tkinter.Menu(menubar,tearoff=0)
    menubar.add_command(label="Logout", command= Logout)

    path =("aait.jpg")
    img = ImageTk.PhotoImage(Image.open(path))
    databutton = tkinter.Button(Home_window, text='Change database password', command = lock)
    databutton.place(x=1000,y=0)
    
    panel = tkinter.Label(Home_window, image = img)
    panel.place(x="-1", y="20")


    Home_window.mainloop()

#mainscreen()
